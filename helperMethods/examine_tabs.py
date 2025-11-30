import pandas as pd
import openpyxl

print("=" * 80)
print("EXAMINING EXCEL SHEET STRUCTURE")
print("=" * 80)

files = {
    'books2016': r'data\books2016.xlsx',
    'books2017': r'data\book2017.xlsx',
    'books2018': r'data\books2018.xlsx',
    'konkurser2019': r'data\konkurser2019.xlsx'
}

for name, filepath in files.items():
    print(f"\n{name}:")
    print("-" * 80)

    # Load workbook to see sheet names
    wb = openpyxl.load_workbook(filepath, read_only=True)
    print(f"Sheet names: {wb.sheetnames}")

    # For each sheet, show basic info
    for sheet_name in wb.sheetnames:
        df = pd.read_excel(filepath, sheet_name=sheet_name)
        print(f"  - Sheet '{sheet_name}': {df.shape[0]} rows × {df.shape[1]} columns")

    wb.close()

print("\n" + "=" * 80)
print("DETAILED KONKURSER2019 ANALYSIS")
print("=" * 80)

# Deep dive into konkurser2019
wb = openpyxl.load_workbook(r'data\konkurser2019.xlsx', read_only=True)

for sheet_name in wb.sheetnames:
    print(f"\n\nSHEET: {sheet_name}")
    print("-" * 80)

    df = pd.read_excel(r'data\konkurser2019.xlsx', sheet_name=sheet_name)
    print(f"Shape: {df.shape[0]} rows × {df.shape[1]} columns")

    # Check if there's a year indicator in the data
    year_cols = [c for c in df.columns if 'år' in str(c).lower() or 'year' in str(c).lower() or 'regnskap' in str(c).lower()]
    if year_cols:
        print(f"Year-related columns: {year_cols}")
        for col in year_cols[:3]:  # Show first 3
            print(f"  {col}: {df[col].iloc[1:6].tolist()}")

    # Show first org number to verify it's real data
    orgnr_col = [c for c in df.columns if 'Orgnr' in str(c)]
    if orgnr_col:
        print(f"Sample Org numbers: {df[orgnr_col[0]].iloc[1:4].tolist()}")

    # Check for accounting data
    accounting_cols = [c for c in df.columns if 'Tall' in str(c) and 'beskrivelse' not in str(c)]
    print(f"Number of accounting metrics: {len(accounting_cols)}")

wb.close()
